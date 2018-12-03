from xlrd import open_workbook
from random import shuffle
from openpyxl import load_workbook
import sys

# Ensure argument passed
if len(sys.argv) != 2:
    sys.exit('Invalid arguments, please specify xls file for scrambling.')

workbook_loc = sys.argv[1]

try:
    wb = open_workbook(workbook_loc)
except IOError:
    sys.exit(str.format('Cannot open {}', workbook_loc))

ws = wb.sheet_by_index(0)
names = ws.col_values(0)
shuffle(names)

wb = load_workbook(workbook_loc)
ws = wb.active
for index, name in enumerate(names):
    ws.cell(row=index+1, column=1).value = name

wb.save(workbook_loc)
